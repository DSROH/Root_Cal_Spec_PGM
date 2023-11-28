import os
import tkinter as tk
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

import tkinter.messagebox as msgbox
import LSI_Solution.Common.get_data as Lget
import LSI_Solution.Common.Cable_check as Lcable
import LSI_Solution.DASEUL.DC_n_IIP2_Cal as Ldciip2
import LSI_Solution.DASEUL.RX_average as Lrx_ave
import LSI_Solution.DASEUL.TX_2g as L2gtx
import LSI_Solution.DASEUL.TX_3g as L3gtx
import LSI_Solution.DASEUL.TX_sub6 as Lsub6tx
import LSI_Solution.DASEUL.ET_3g as L3get
import LSI_Solution.DASEUL.ET_Sub6 as Lsub6et
import LSI_Solution.DASEUL.FBRX as Lfbrx
import LSI_Solution.DASEUL.APT as Lapt
import LSI_Solution.Common.Function as func
# import LSI_Solution.DASEUL.Drawing as Ldraw


def rfic_gain_average(df_rfic_gain, Save_data, text_area):
    df_rfic_gain_Value = df_rfic_gain.iloc[:, 1:2].astype(float)
    df_rfic_gain_Item = df_rfic_gain["Test Conditions"].str.split("_", expand=True)
    #     # 의미없는 컬럼 삭제
    df_rfic_gain_Item.drop(columns=[3, 4], inplace=True)
    # # groupby 실행을 위한 컬럼명 변경
    df_rfic_gain_Item.columns = ["RAT", "Band", "Path", "Index"]
    df_rfic_gain = pd.merge(df_rfic_gain_Item, df_rfic_gain_Value, left_index=True, right_index=True)
    df_rfic_gain_Mean = round(df_rfic_gain.groupby(["RAT", "Band", "Path", "Index"], sort=False).mean(), 1)
    # df_rfic_gain_Data = round(df_rfic_gain.groupby(["RAT", "Band", "Path", "Index"], sort=False).agg(["mean", "max", "min"]), 1)
    df_rfic_gain_Mean = pd.concat([df_rfic_gain_Mean, round(df_rfic_gain_Mean.mean(axis=1), 1).rename("Average")], axis=1)

    if Save_data:
        filename = "Excel_RFIC_Gain.xlsx"
        with pd.ExcelWriter(filename) as writer:
            df_rfic_gain_Mean.to_excel(writer, sheet_name="RFIC_Gain_Mean")
        func.WB_Format(filename, 2, 4, 0, text_area)

    return df_rfic_gain_Mean["Average"]


def get_data(list_file, Select_op, Selected_Option, Save_data_var, debug_var, raw_data_var, get_data_var, text_area):
    if list_file.size() == 0:
        msgbox.showwarning("경고", "Cal log 파일(*.csv)을 추가하세요")
        return

    text_area.delete("1.0", "end")
    text_area.insert(tk.END, "=" * 85)
    text_area.insert(tk.END, "\n")
    Result_Meas = pd.DataFrame()
    Result_Code = pd.DataFrame()

    Strt_dir = os.path.dirname(list_file.get(0))

    if Save_data_var.get() or raw_data_var.get() or get_data_var.get():
        today = datetime.today().strftime("%Y_%m_%d")
        start_T = datetime.today().strftime("%Y%m%d_%H%M")
        save_dir = Strt_dir + "\\Exported_Data\\" + today + "\\"
        func.createDirectory(save_dir)
    else:
        save_dir = Strt_dir

    os.chdir(save_dir)
    file_list = list_file.get(0, tk.END)

    if Select_op == "Daseul":
        Total_count = 0
        for file in file_list:
            Print_fname = os.path.basename(file)
            Print_Word = "Collecting Data"
            text_area.insert(tk.END, f"{Print_Word:<25}     | {Print_fname:<38} | ")
            text_area.see(tk.END)
            # Import Data
            # my_cols = [str(i) for i in range(30)]  # create some col names
            my_cols = [
                "Test Conditions",
                "Meas",
                "Lower Limits",
                "Upper Limits",
                "P_F",
                "Sec",
                "Code",
                "Code LSL",
                "Code USL",
                "Meas Fine",
                "Code Fine",
            ]
            df_Data = pd.read_csv(file, sep="\t|,", names=my_cols, header=None, engine="python", encoding="utf-8")
            text_area.insert(tk.END, f"Done\n")
            text_area.see(tk.END)
            # df_null = df_Data[df_Data["Test Conditions"].isnull()].index
            # df_Data.drop(df_null, inplace=True)
            # EPT가 Auto Sweep 이라서 캘로그마다 실행갯수가 다르다 -> 추출 후 Drop
            # Print_Word = "Drop RFIC Gain Cal Data"
            # text_area.insert(tk.END, f"{Print_Word:<25}     | {Print_fname:<38} | ")
            # text_area.see(tk.END)
            df_RFIC_gain = df_Data[df_Data["Test Conditions"].str.contains("_RFIC_")]
            df_RFIC_gain.reset_index(drop=True, inplace=True)
            df_Data.drop(df_Data[df_Data["Test Conditions"].str.contains("_RFIC_")].index, inplace=True)
            df_Data.reset_index(drop=True, inplace=True)
            # text_area.insert(tk.END, f"Done\n")
            # text_area.see(tk.END)

            count = df_Data[df_Data["Test Conditions"].str.contains("// << UartSwitchToCP >>")].shape[0]

            df_Meas = df_Data[["Test Conditions", "Meas"]].reset_index(drop=True)
            df_Code = df_Data[["Test Conditions", "Code"]].reset_index(drop=True)

            Start_Meas = df_Meas.index[df_Meas["Test Conditions"].str.contains("// << CMC_RFCableCheck >>")].tolist()
            Stop_Meas = df_Meas.index[df_Meas["Test Conditions"].str.contains("// << H/W Version Write >>")].tolist()
            Start_Code = df_Code.index[df_Code["Test Conditions"].str.contains("// << WCDMA Tx DC Calibration >>")].tolist()
            Stop_Code = df_Code.index[df_Code["Test Conditions"].str.contains("// << SLSI_CAL_HSPA_POST_V3 >>")].tolist()

            size_missmatch = False if (len(Start_Meas) == len(Stop_Meas)) else True

            Re_test_Count = 0
            for i in range(count):
                Re_test = False
                if size_missmatch:
                    Subset_Meas = df_Meas[Start_Meas[i] : Stop_Meas[i - Re_test_Count]]
                    Subset_Code = df_Code[Start_Code[i] : Stop_Code[i - Re_test_Count]]
                else:
                    Subset_Meas = df_Meas[Start_Meas[i] : Stop_Meas[i]]
                    Subset_Code = df_Code[Start_Code[i] : Stop_Code[i]]

                Re_test = any(Subset_Meas["Test Conditions"].str.contains("// Retry"))
                if Re_test:
                    Re_test_Count += 1
                    continue

                Subset_Meas = Subset_Meas.dropna().reset_index(drop=True)
                Subset_Code = Subset_Code.dropna().reset_index(drop=True)

                if debug_var.get():
                    # Debug only
                    print(f"Subset_Meas Size = {len(Subset_Meas)}")
                    print(f"Subset_Code Size = {len(Subset_Code)}")
                    Subset_Meas.to_excel(f"Subset_Meas_{i+1}.xlsx")
                    Subset_Code.to_excel(f"Subset_Code_{i+1}.xlsx")

                if Re_test & (len(Subset_Meas) > 20751):
                    del Start_Meas[i]
                    if Start_Meas[i] > Stop_Meas[i]:
                        del Stop_Meas[i]
                    Subset_Meas = df_Meas[Start_Meas[i] : Stop_Meas[i]]

                    if debug_var.get():
                        print(f"Deleted Meas = {Start_Meas} to {Stop_Meas}")

                if i == 0:
                    Result_Meas["Test Conditions"] = Subset_Meas["Test Conditions"]
                    Result_Code["Test Conditions"] = Subset_Code["Test Conditions"]
                    # groupby 실행 시 숫자가 아닌 열은 자동 생략 (ommiting nuisance) 처리됨, Pandas 차기버전부터 오류로 처리
                    # -> astype으로 Object를 float으로 변경
                    Result_Meas = pd.concat([Result_Meas, Subset_Meas["Meas"].astype(float).rename(f"Meas_1")], axis=1)
                    Result_Code = pd.concat([Result_Code, Subset_Code["Code"].astype(float).rename(f"Code_1")], axis=1)

                else:
                    # groupby 실행 시 숫자가 아닌 열은 자동 생략 (ommiting nuisance) 처리됨, Pandas 차기버전부터 오류로 처리
                    # -> astype으로 Object를 float으로 변경
                    Result_Meas = pd.concat(
                        [Result_Meas, Subset_Meas["Meas"].astype(float).rename(f"Meas_{Total_count+i+1}")], axis=1
                    )
                    Result_Code = pd.concat(
                        [Result_Code, Subset_Code["Code"].astype(float).rename(f"Code_{Total_count+i+1}")], axis=1
                    )

                text_area.insert(tk.END, f"Data Count = {Total_count+i+1}\n")
                text_area.see(tk.END)

            Total_count += i

        Result_Meas.dropna(how="all", axis=1, inplace=True)
        Result_Code.dropna(how="all", axis=1, inplace=True)

        text_area.insert(tk.END, "=" * 85)
        text_area.insert(tk.END, "\n")
        text_area.see(tk.END)

        if raw_data_var.get():
            filename = "Result_Meas.xlsx"
            with pd.ExcelWriter(filename) as writer:
                Result_Meas.to_excel(writer, sheet_name="Meas_Full_Data")
            wb = load_workbook(filename)
            ws = wb.sheetnames
            for sheet in ws:
                col_max = wb[sheet].max_column
                row_max = wb[sheet].max_row
                text_area.insert(tk.END, f"{sheet:<30}| Col = {col_max:<5}, Row = {row_max:<5} | ")
                text_area.see(tk.END)
            text_area.insert(tk.END, f"Done\n")
            text_area.see(tk.END)

            filename = "Result_Code.xlsx"
            with pd.ExcelWriter(filename) as writer:
                Result_Code.to_excel(writer, sheet_name="Code_Full_Data")
            wb = load_workbook(filename)
            ws = wb.sheetnames
            for sheet in ws:
                col_max = wb[sheet].max_column
                row_max = wb[sheet].max_row
                text_area.insert(tk.END, f"{sheet:<30}| Col = {col_max:<5}, Row = {row_max:<5} | ")
                text_area.see(tk.END)
            text_area.insert(tk.END, f"Done\n")
            text_area.see(tk.END)

        if get_data_var.get():
            Save_data = True
        else:
            Save_data = Save_data_var.get()

        DC_Cal_Mean, IIP2_Cal_Mean, DC_Cal_3G, DC_Cal_2G, DC_Cal_NR = Ldciip2.daseul_dc_iip2(Result_Code, Save_data, text_area)
        CableCheck = Lcable.daseul_cable_average(Result_Meas, Save_data, text_area)
        RFIC_gain = Lget.rfic_gain_average(df_RFIC_gain, Save_data, text_area)
        TxP_Channel_comp_ave = L3gtx.TxP_3g_channel_comp_pa_mid(Result_Meas, Save_data, text_area)
        (
            FBRX_Gain_Meas_3G,
            FBRX_Gain_Meas_sub6,
            FBRX_Gain_Code_3G,
            FBRX_Gain_Code_sub6,
            FBRX_Freq_Meas_3G,
            FBRX_Freq_Meas_3G_Max,
            FBRX_Freq_Meas_3G_Min,
            FBRX_Freq_Meas_sub6,
            FBRX_Freq_Meas_sub6_Max,
            FBRX_Freq_Meas_sub6_Min,
            FBRX_Freq_Code_sub6,
            FBRX_Freq_Code_sub6_Max,
            FBRX_Freq_Code_sub6_Min,
        ) = Lfbrx.fbrx_average(Result_Meas, Result_Code, Save_data, text_area)
        PRX_Gain_2G, Ripple_2G, RXGain_3G, RXComp_3G, RXGain_sub6, RXRSRP_sub6, RXComp_sub6 = Lrx_ave.daseul_rx_average(
            Result_Meas, Save_data, text_area
        )
        GMSK_Mean, GMSK_TXL_Mean, GMSK_Code_Mean = L2gtx.gsm_tx_gmsk_average(Result_Meas, Result_Code, Save_data, text_area)
        EPSK_Mean, EPSK_TXL_Mean, EPSK_Code_Mean = L2gtx.gsm_tx_edge_average(Result_Meas, Result_Code, Save_data, text_area)

        (
            ETSAPT_3G_Psat_Ave,
            ETSAPT_3G_Psat_Max,
            ETSAPT_3G_Psat_Min,
            ETSAPT_3G_Power_Ave,
            ETSAPT_3G_Power_Max,
            ETSAPT_3G_Power_Min,
        ) = L3get.Et_3g_average(Result_Meas, Save_data, text_area)

        thermistor = Lsub6tx.therm_average(Result_Code, Save_data, text_area)

        APT_3G_Ave, APT_Sub6_Ave, APT_Sub6_Max, APT_Sub6_Min = Lapt.apt_average(Result_Meas, Save_data, text_area)

        (
            ETSAPT_sub6_Psat_Ave,
            ETSAPT_sub6_Psat_Max,
            ETSAPT_sub6_Psat_Min,
            ETSAPT_sub6_Freq_Ave,
            ETSAPT_sub6_Freq_Max,
            ETSAPT_sub6_Freq_Min,
            ETSAPT_sub6_Pgain_Ave,
            ETSAPT_sub6_Pgain_Max,
            ETSAPT_sub6_Pgain_Min,
            ETSAPT_sub6_Power_Ave,
            ETSAPT_sub6_Power_Max,
            ETSAPT_sub6_Power_Min,
        ) = Lsub6et.sub6_et_average(Result_Meas, Save_data, text_area)
        BW_Cal = Lsub6tx.sub6_bw_cal_average(Result_Meas, Save_data, text_area)

        # ! Plotly
        # Ldraw.Drawing_dc_cal(DC_Cal_Mean)

        if get_data_var.get() & (Selected_Option == 4):
            msgbox.showwarning("Message", "작업 완료")
            get_data_var.set(False)
        else:
            return (
                CableCheck,
                RFIC_gain,
                TxP_Channel_comp_ave,
                FBRX_Gain_Meas_3G,
                FBRX_Gain_Meas_sub6,
                FBRX_Gain_Code_3G,
                FBRX_Gain_Code_sub6,
                FBRX_Freq_Meas_3G,
                FBRX_Freq_Meas_3G_Max,
                FBRX_Freq_Meas_3G_Min,
                FBRX_Freq_Meas_sub6,
                FBRX_Freq_Meas_sub6_Max,
                FBRX_Freq_Meas_sub6_Min,
                FBRX_Freq_Code_sub6,
                FBRX_Freq_Code_sub6_Max,
                FBRX_Freq_Code_sub6_Min,
                PRX_Gain_2G,
                Ripple_2G,
                RXGain_3G,
                RXComp_3G,
                RXGain_sub6,
                RXRSRP_sub6,
                RXComp_sub6,
                GMSK_Mean,
                GMSK_TXL_Mean,
                GMSK_Code_Mean,
                EPSK_Mean,
                EPSK_TXL_Mean,
                EPSK_Code_Mean,
                ETSAPT_3G_Psat_Ave,
                ETSAPT_3G_Psat_Max,
                ETSAPT_3G_Psat_Min,
                ETSAPT_3G_Power_Ave,
                ETSAPT_3G_Power_Max,
                ETSAPT_3G_Power_Min,
                thermistor,
                APT_3G_Ave,
                APT_Sub6_Ave,
                APT_Sub6_Max,
                APT_Sub6_Min,
                ETSAPT_sub6_Psat_Ave,
                ETSAPT_sub6_Psat_Max,
                ETSAPT_sub6_Psat_Min,
                ETSAPT_sub6_Freq_Ave,
                ETSAPT_sub6_Freq_Max,
                ETSAPT_sub6_Freq_Min,
                ETSAPT_sub6_Pgain_Ave,
                ETSAPT_sub6_Pgain_Max,
                ETSAPT_sub6_Pgain_Min,
                ETSAPT_sub6_Power_Ave,
                ETSAPT_sub6_Power_Max,
                ETSAPT_sub6_Power_Min,
                BW_Cal,
            )

    else:  # MTM
        for count, file in enumerate(file_list, start=1):
            Print_fname = os.path.basename(file)
            Print_Word = "Collecting Data"
            text_area.insert(tk.END, f"{Print_Word:<30}|\t{Print_fname}\t|\t")

            my_cols = [str(i) for i in range(20)]  # create some col names
            df_Data = pd.read_csv(file, sep="\t|,|=", names=my_cols, header=None, engine="python")
            text_area.insert(tk.END, f"Done\n")

            Print_Word = "Drop Null Data"
            text_area.insert(tk.END, f"{Print_Word:<30}|\t{Print_fname}\t|\t")
            df_Data = df_Data.iloc[2:, :3]
            df_null = df_Data[df_Data["0"].isnull()].index
            df_Data.drop(df_null, inplace=True)
            text_area.insert(tk.END, f"Done\n")

            Print_Word = "RFIC Gain Cal Data"
            text_area.insert(tk.END, f"{Print_Word:<30}|\t{Print_fname}\t|\t")
            text_area.see(tk.END)

            df_RFIC_gain = df_Data[df_Data["1"].str.contains("RFIC Gain")]
            RFIC_Gain_index = df_Data[df_Data["1"].str.contains("RFIC Gain")].index.tolist()
            HSPA_Modul_index = df_Data[df_Data["1"].str.contains(" Modulation FBRx ")].index.tolist()
            RFIC_Gain_index.extend(HSPA_Modul_index)

            Subset_Meas = df_Data.drop(RFIC_Gain_index).reset_index(drop=True)
            Subset_Code = df_Data[df_Data["1"].str.contains(" value")].dropna().reset_index(drop=True)
            text_area.insert(tk.END, f"Done\n")

            Subset_Meas.columns = ["Band", "Item", "Meas"]
            Subset_Code.columns = ["Band", "Item", "Code"]
            # groupby 실행 시 숫자가 아닌 열은 자동 생략 (ommiting nuisance) 처리됨, Pandas 차기버전부터 오류로 처리
            # -> astype으로 Object를 float으로 변경
            if count == 1:
                Result_Meas[["Band", "Item"]] = Subset_Meas[["Band", "Item"]]
                Result_Code[["Band", "Item"]] = Subset_Code[["Band", "Item"]]
                Result_Meas["Meas_1"] = Subset_Meas["Meas"].astype(float)
                Result_Code["Code_1"] = Subset_Code["Code"].astype(float)
            else:
                Result_Meas[f"Meas_{count}"] = Subset_Meas["Meas"].astype(float)
                Result_Code[f"Code_{count}"] = Subset_Code["Code"].astype(float)

        Result_Meas.dropna(how="all", axis=1, inplace=True)
        Result_Code.dropna(how="all", axis=1, inplace=True)

        text_area.insert(tk.END, "=" * 85)
        text_area.insert(tk.END, "\n")
        text_area.see(tk.END)

        return Result_Meas
