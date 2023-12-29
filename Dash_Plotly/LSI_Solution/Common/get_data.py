import os
import threading
import tkinter as tk
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

import tkinter.messagebox as msgbox
import LSI_Solution.Common.get_data as Lget
import LSI_Solution.Common.Cable_check as Lcable
import LSI_Solution.DASEUL.DC_n_IIP2_Cal as Ldciip2
import LSI_Solution.DASEUL.RX_ave as Lrx_ave
import LSI_Solution.DASEUL.TX_2g as L2gtx
import LSI_Solution.DASEUL.TX_3g as L3gtx
import LSI_Solution.DASEUL.TX_sub6 as Lsub6tx
import LSI_Solution.DASEUL.ET_3g as L3get
import LSI_Solution.DASEUL.ET_Sub6 as Lsub6et
import LSI_Solution.DASEUL.FBRX as Lfbrx
import LSI_Solution.DASEUL.APT as Lapt
import LSI_Solution.Common.Function as func
from LSI_Solution.pages.dash_app import MultiPageApp
import traceback as tb


def Rfic_gain_ave(df_rfic_gain, Save_data, text_area):
    df_rfic_gain_Value = df_rfic_gain.iloc[:, 1:2].astype(float)
    df_rfic_gain_Item = df_rfic_gain["Test Conditions"].str.split("_", expand=True)
    #     # 의미없는 컬럼 삭제
    df_rfic_gain_Item.drop(columns=[3, 4], inplace=True)
    # # groupby 실행을 위한 컬럼명 변경
    df_rfic_gain_Item.columns = ["RAT", "Band", "Path", "Index"]
    df_rfic_gain = pd.merge(df_rfic_gain_Item, df_rfic_gain_Value, left_index=True, right_index=True)
    df_rfic_gain_Mean = round(df_rfic_gain.groupby(["RAT", "Band", "Path", "Index"], sort=False).mean(), 1)
    # df_rfic_gain_Data = round(df_rfic_gain.groupby(["RAT", "Band", "Path", "Index"], sort=False).agg(["mean", "max", "min"]), 1)
    df_rfic_gain_Mean = pd.concat([df_rfic_gain_Mean, round(df_rfic_gain_Mean.mean(axis=1), 1).rename("Average")], axis=1)

    if Save_data:
        filename = "Excel_RFIC_Gain.xlsx"
        with pd.ExcelWriter(filename) as writer:
            df_rfic_gain_Mean.to_excel(writer, sheet_name="RFIC_Gain_Mean")
        func.WB_Format(filename, 2, 4, 0, text_area)

    return df_rfic_gain_Mean["Average"]


def get_data(list_file, Select_op, Selected_Option, Save_data_var, debug_var, raw_data_var, get_data_var, text_area):
    try:
        if list_file.size() == 0:
            msgbox.showwarning("경고", "Cal log 파일(*.csv)을 추가하세요")
            return

        text_area.delete("1.0", "end")
        text_area.insert(tk.END, "=" * 85)
        text_area.insert(tk.END, "\n")
        Result_Meas = pd.DataFrame()
        Result_Code = pd.DataFrame()

        Strt_dir = os.path.dirname(list_file.get(0))

        if Save_data_var.get() or raw_data_var.get() or get_data_var.get():
            today = datetime.today().strftime("%Y_%m_%d")
            start_T = datetime.today().strftime("%Y%m%d_%H%M")
            save_dir = Strt_dir + "\\Exported_Data\\" + today + "\\"
            func.createDirectory(save_dir)
        else:
            save_dir = Strt_dir

        os.chdir(save_dir)
        file_list = list_file.get(0, tk.END)

        if Select_op == "Daseul":
            Total_count = 0
            Loop_count = 0
            for file in file_list:
                Print_fname = os.path.basename(file)
                Print_Word = "Collecting Data"
                text_area.insert(tk.END, f"{Print_Word:<25}     | {Print_fname:<38} | ")
                text_area.see(tk.END)
                # Import Data
                # my_cols = [str(i) for i in range(30)]  # create some col names
                my_cols = [
                    "Test Conditions",
                    "Meas",
                    "Lower Limits",
                    "Upper Limits",
                    "P_F",
                    "Sec",
                    "Code",
                    "Code LSL",
                    "Code USL",
                    "Meas Fine",
                    "Code Fine",
                ]
                df_Data = pd.read_csv(file, sep="\t|,", names=my_cols, header=None, engine="python", encoding="utf-8")
                text_area.insert(tk.END, f"Done\n")
                text_area.see(tk.END)
                # df_null = df_Data[df_Data["Test Conditions"].isnull()].index
                # df_Data.drop(df_null, inplace=True)
                # EPT가 Auto Sweep 이라서 캘로그마다 실행갯수가 다르다 -> 추출 후 Drop
                # Print_Word = "Drop RFIC Gain Cal Data"
                # text_area.insert(tk.END, f"{Print_Word:<25}     | {Print_fname:<38} | ")
                # text_area.see(tk.END)
                df_RFIC_gain = df_Data[df_Data["Test Conditions"].str.contains("_RFIC_")]
                df_RFIC_gain.reset_index(drop=True, inplace=True)
                df_Data.drop(df_Data[df_Data["Test Conditions"].str.contains("_RFIC_")].index, inplace=True)
                df_Data.reset_index(drop=True, inplace=True)
                # text_area.insert(tk.END, f"Done\n")
                # text_area.see(tk.END)

                count = df_Data[df_Data["Test Conditions"].str.contains("// << UartSwitchToCP >>")].shape[0]

                df_Meas = df_Data[["Test Conditions", "Meas"]].reset_index(drop=True)
                df_Code = df_Data[["Test Conditions", "Code"]].reset_index(drop=True)

                Start_Meas = df_Meas.index[df_Meas["Test Conditions"].str.contains("// << CMC_RFCableCheck >>")].tolist()
                Stop_Meas = df_Meas.index[df_Meas["Test Conditions"].str.contains("// << H/W Version Write >>")].tolist()
                Start_Code = df_Code.index[df_Code["Test Conditions"].str.contains("// << WCDMA Tx DC Calibration >>")].tolist()
                Stop_Code = df_Code.index[df_Code["Test Conditions"].str.contains("// << SLSI_CAL_HSPA_POST_V3 >>")].tolist()

                size_missmatch = False if (len(Start_Meas) == len(Stop_Meas)) else True
                Re_test_Count = 0
                for i in range(1, count + 1, 1):
                    Total_count = Loop_count + i
                    Re_test = False
                    # ? 오류있는 결과들을 삭제해서 현재 카운트가 Start_Meas의 인덱스보다 커질경우 for문 탈출
                    if i > len(Start_Meas):
                        break

                    if size_missmatch:
                        Subset_Meas = df_Meas[Start_Meas[i - 1] : Stop_Meas[i - 1 - Re_test_Count]]
                        Subset_Code = df_Code[Start_Code[i - 1] : Stop_Code[i - 1 - Re_test_Count]]
                    else:
                        Subset_Meas = df_Meas[Start_Meas[i - 1] : Stop_Meas[i - 1]]
                        Subset_Code = df_Code[Start_Code[i - 1] : Stop_Code[i - 1]]

                    Re_test = any(Subset_Meas["Test Conditions"].str.contains("// Retry")) or any(
                        Subset_Code["Test Conditions"].str.contains("// Retry")
                    )
                    if Re_test:
                        Re_test_Count += 1
                        print(f"Data Count = {Total_count} Deleted. Detected : // Retry")
                        text_area.insert(tk.END, f"Data Count = {Total_count} Deleted. Detected : // Retry\n")
                        text_area.see(tk.END)
                        continue

                    Subset_Meas = Subset_Meas.dropna().reset_index(drop=True)
                    Subset_Code = Subset_Code.dropna().reset_index(drop=True)

                    if debug_var.get():
                        # Debug only
                        print(f"Subset_Meas_{Total_count} Size = {len(Subset_Meas)}")
                        print(f"Subset_Code_{Total_count} Size = {len(Subset_Code)}")
                        Subset_Meas.to_excel(f"Subset_Meas_{Total_count}.xlsx")
                        Subset_Code.to_excel(f"Subset_Code_{Total_count}.xlsx")

                    if len(Subset_Meas) > 5141 or len(Subset_Code) > 1311:
                        del Start_Meas[i - 1]
                        del Start_Code[i - 1]
                        if Start_Meas[i - 1] > Stop_Meas[i - 1]:
                            del Stop_Meas[i - 1]
                        if Start_Code[i - 1] > Stop_Code[i - 1]:
                            del Stop_Code[i - 1]
                        Subset_Meas = df_Meas[Start_Meas[i - 1] : Stop_Meas[i - 1]]
                        Subset_Code = df_Meas[Start_Code[i - 1] : Stop_Code[i - 1]]
                        text_area.insert(tk.END, f"Data Count = {Total_count} Deleted. Length Over\n")
                        text_area.see(tk.END)
                        print(f"Data Count = {Total_count} Deleted. Length Over")
                        continue

                    if Total_count == 1:
                        Result_Meas["Test Conditions"] = Subset_Meas["Test Conditions"]
                        Result_Code["Test Conditions"] = Subset_Code["Test Conditions"]
                        # groupby 실행 시 숫자가 아닌 열은 자동 생략 (ommiting nuisance) 처리됨, Pandas 차기버전부터 오류로 처리
                        # -> astype으로 Object를 float으로 변경
                        Result_Meas = pd.concat(
                            [
                                Result_Meas,
                                Subset_Meas["Meas"].astype(float).rename(f"Meas_1"),
                            ],
                            axis=1,
                        )
                        Result_Code = pd.concat(
                            [
                                Result_Code,
                                Subset_Code["Code"].astype(float).rename(f"Code_1"),
                            ],
                            axis=1,
                        )

                    else:
                        # groupby 실행 시 숫자가 아닌 열은 자동 생략 (ommiting nuisance) 처리됨, Pandas 차기버전부터 오류로 처리
                        # -> astype으로 Object를 float으로 변경
                        Result_Meas = pd.concat(
                            [
                                Result_Meas,
                                Subset_Meas["Meas"].astype(float).rename(f"Meas_{Total_count}"),
                            ],
                            axis=1,
                        )
                        Result_Code = pd.concat(
                            [
                                Result_Code,
                                Subset_Code["Code"].astype(float).rename(f"Code_{Total_count}"),
                            ],
                            axis=1,
                        )

                    text_area.insert(tk.END, f"Data Count = {Total_count}\n")
                    text_area.see(tk.END)

                Loop_count += i

            Result_Meas.dropna(how="all", axis=0, inplace=True)
            Result_Meas.dropna(how="all", axis=1, inplace=True)
            Result_Code.dropna(how="all", axis=0, inplace=True)
            Result_Code.dropna(how="all", axis=1, inplace=True)

            text_area.insert(tk.END, "=" * 85)
            text_area.insert(tk.END, "\n")
            text_area.see(tk.END)

            if raw_data_var.get():
                filename = "Result_Meas.xlsx"
                with pd.ExcelWriter(filename) as writer:
                    Result_Meas.to_excel(writer, sheet_name="Meas_Full_Data")
                wb = load_workbook(filename)
                ws = wb.sheetnames
                for sheet in ws:
                    col_max = wb[sheet].max_column
                    row_max = wb[sheet].max_row
                    text_area.insert(tk.END, f"{sheet:<30}| Col = {col_max:<5}, Row = {row_max:<5} | ")
                    text_area.see(tk.END)
                text_area.insert(tk.END, f"Done\n")
                text_area.see(tk.END)

                filename = "Result_Code.xlsx"
                with pd.ExcelWriter(filename) as writer:
                    Result_Code.to_excel(writer, sheet_name="Code_Full_Data")
                wb = load_workbook(filename)
                ws = wb.sheetnames
                for sheet in ws:
                    col_max = wb[sheet].max_column
                    row_max = wb[sheet].max_row
                    text_area.insert(tk.END, f"{sheet:<30}| Col = {col_max:<5}, Row = {row_max:<5} | ")
                    text_area.see(tk.END)
                text_area.insert(tk.END, f"Done\n")
                text_area.see(tk.END)

            if get_data_var.get():
                Save_data = True
            else:
                Save_data = Save_data_var.get()

            DC_Cal_Mean, IIP2_Cal_Mean = Ldciip2.Daseul_dc_iip2(Result_Code, Save_data, text_area)
            CableCheck = Lcable.Daseul_cable_ave(Result_Meas, Save_data, text_area)
            RFIC_gain = Lget.Rfic_gain_ave(df_RFIC_gain, Save_data, text_area)
            TxPCh_Comp_3g = L3gtx.Txp_cc_3g(Result_Meas, Save_data, text_area)
            (
                FBRX_Gain_Meas_3G,
                FBRX_Gain_Code_3G,
                FBRX_Freq_Meas_3G,
                FBRX_Freq_Meas_3G_CH,
                FBRX_Gain_Meas_sub6,
                FBRX_Gain_Code_sub6,
                FBRX_Freq_Meas_sub6,
                FBRX_Freq_Code_sub6,
            ) = Lfbrx.Fbrx_ave(Result_Meas, Result_Code, Save_data, text_area)
            (
                PRX_Gain_2G,
                Ripple_2G,
                RXGain_3G,
                RXComp_3G,
                RXGain_sub6,
                RXRSRP_sub6,
                RXComp_sub6,
                RXMixer_sub6,
            ) = Lrx_ave.Daseul_rx_ave(Result_Meas, Save_data, text_area)
            GMSK_Mean, GMSK_TxL_Mean, GMSK_Code_Mean = L2gtx.Gsm_tx_gmsk_ave(Result_Meas, Result_Code, Save_data, text_area)
            EPSK_Mean, EPSK_TxL_Mean, EPSK_Code_Mean = L2gtx.Gsm_tx_edge_ave(Result_Meas, Result_Code, Save_data, text_area)
            ET_3G_Psat, ET_3G_Power = L3get.Et_3g_ave(Result_Meas, Save_data, text_area)
            thermistor = Lsub6tx.Therm_ave(Result_Code, Save_data, text_area)
            APT_Meas_3G, APT_Meas_Sub6 = Lapt.Apt_ave(Result_Meas, Save_data, text_area)
            ET_sub6_Psat, ET_sub6_Pgain, ET_sub6_Power, ET_sub6_Freq = Lsub6et.Sub6_et_ave(Result_Meas, Save_data, text_area)
            BW_Cal = Lsub6tx.Sub6_bw_cal_ave(Result_Meas, Save_data, text_area)

            dict_cf = {"txdc": DC_Cal_Mean, "iip2": IIP2_Cal_Mean, "cable": CableCheck}

            dict_2g = {
                "rx_gain": PRX_Gain_2G,
                "rx_ripp": Ripple_2G,
                "gmsk": GMSK_Mean,
                "gmsk_txl": GMSK_TxL_Mean,
                "gmsk_code": GMSK_Code_Mean,
                "epsk": EPSK_Mean,
                "epsk_txl": EPSK_TxL_Mean,
                "epsk_code": EPSK_Code_Mean,
            }

            dict_3g = {
                "rx_gain": RXGain_3G,
                "rx_comp": RXComp_3G,
                "fbrx_gm": FBRX_Gain_Meas_3G,
                "fbrx_gc": FBRX_Gain_Code_3G,
                "fbrx_fm": FBRX_Freq_Meas_3G,
                "FBRX_FM_Ch": FBRX_Freq_Meas_3G_CH,
                "apt_meas": APT_Meas_3G,
                "txp_cc": TxPCh_Comp_3g,
                "et_psat": ET_3G_Psat,
                "et_power": ET_3G_Power,
            }

            dict_nr = {
                "rx_gain": RXGain_sub6,
                "rx_rsrp": RXRSRP_sub6,
                "rx_comp": RXComp_sub6,
                "rx_mix": RXMixer_sub6,
                "fbrx_gm": FBRX_Gain_Meas_sub6,
                "fbrx_gc": FBRX_Gain_Code_sub6,
                "fbrx_fm": FBRX_Freq_Meas_sub6,
                "fbrx_fc": FBRX_Freq_Code_sub6,
                "apt_meas": APT_Meas_Sub6,
                "therm": thermistor,
                "et_psat": ET_sub6_Psat,
                "et_pgain": ET_sub6_Pgain,
                "et_power": ET_sub6_Power,
                "et_freq": ET_sub6_Freq,
                "bw_cal": BW_Cal,
            }

            # ** ============================== Plotly ==============================

            # Ldraw.Drawing_dash(DC_Cal_Mean)
            threading.Thread(target=MultiPageApp, args=(dict_cf, dict_2g, dict_3g, dict_nr)).start()

            # ** ============================== Plotly ==============================

            if get_data_var.get() & (Selected_Option == 4):
                msgbox.showwarning("Message", "작업 완료")
                get_data_var.set(False)
            else:
                return dict_cf, dict_2g, dict_3g, dict_nr

        else:  # MTM
            for count, file in enumerate(file_list, start=1):
                Print_fname = os.path.basename(file)
                Print_Word = "Collecting Data"
                text_area.insert(tk.END, f"{Print_Word:<30}|\t{Print_fname}\t|\t")

                my_cols = [str(i) for i in range(20)]  # create some col names
                df_Data = pd.read_csv(file, sep="\t|,|=", names=my_cols, header=None, engine="python")
                text_area.insert(tk.END, f"Done\n")

                Print_Word = "Drop Null Data"
                text_area.insert(tk.END, f"{Print_Word:<30}|\t{Print_fname}\t|\t")
                df_Data = df_Data.iloc[2:, :3]
                df_null = df_Data[df_Data["0"].isnull()].index
                df_Data.drop(df_null, inplace=True)
                text_area.insert(tk.END, f"Done\n")

                Print_Word = "RFIC Gain Cal Data"
                text_area.insert(tk.END, f"{Print_Word:<30}|\t{Print_fname}\t|\t")
                text_area.see(tk.END)

                df_RFIC_gain = df_Data[df_Data["1"].str.contains("RFIC Gain")]
                RFIC_Gain_index = df_Data[df_Data["1"].str.contains("RFIC Gain")].index.tolist()
                HSPA_Modul_index = df_Data[df_Data["1"].str.contains(" Modulation FBRx ")].index.tolist()
                RFIC_Gain_index.extend(HSPA_Modul_index)

                Subset_Meas = df_Data.drop(RFIC_Gain_index).reset_index(drop=True)
                Subset_Code = df_Data[df_Data["1"].str.contains(" value")].dropna().reset_index(drop=True)
                text_area.insert(tk.END, f"Done\n")

                Subset_Meas.columns = ["Band", "Item", "Meas"]
                Subset_Code.columns = ["Band", "Item", "Code"]
                # groupby 실행 시 숫자가 아닌 열은 자동 생략 (ommiting nuisance) 처리됨, Pandas 차기버전부터 오류로 처리
                # -> astype으로 Object를 float으로 변경
                if count == 1:
                    Result_Meas[["Band", "Item"]] = Subset_Meas[["Band", "Item"]]
                    Result_Code[["Band", "Item"]] = Subset_Code[["Band", "Item"]]
                    Result_Meas["Meas_1"] = Subset_Meas["Meas"].astype(float)
                    Result_Code["Code_1"] = Subset_Code["Code"].astype(float)
                else:
                    Result_Meas[f"Meas_{count}"] = Subset_Meas["Meas"].astype(float)
                    Result_Code[f"Code_{count}"] = Subset_Code["Code"].astype(float)

            Result_Meas.dropna(how="all", axis=0, inplace=True)
            Result_Meas.dropna(how="all", axis=1, inplace=True)

            text_area.insert(tk.END, "=" * 85)
            text_area.insert(tk.END, "\n")
            text_area.see(tk.END)

            return Result_Meas

    except Exception:
        msgbox.showwarning("ERROR", tb.format_exc())
