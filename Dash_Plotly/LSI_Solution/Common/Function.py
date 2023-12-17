import os
import tkinter as tk
from tkinter import filedialog

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.numbers import builtin_format_code


def Common_save_Excel(filename, tab1, tab2):
    # Save Data to Excel
    Tabname = filename.replace("Excel_", "")
    Tabname = f"{os.path.splitext(Tabname)[0]}"
    with pd.ExcelWriter(filename) as writer:
        tab1.to_excel(writer, sheet_name=f"{Tabname}_Mean")
        tab2.to_excel(writer, sheet_name=f"{Tabname}_Data")


# def Common_save_Excel(filename, *args):
#     # Save Data to Excel
#     Tabname = filename.replace("Excel_", "")
#     Tabname = f"{os.path.splitext(Tabname)[0]}"
#     with pd.ExcelWriter(filename) as writer:
#         for i in args:
#             i.to_excel(writer, sheet_name=f"{Tabname}")


def createDirectory(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print("\tError: Failed to create directory.")


def WB_Format(filename, i, j, k, text_area):
    font_style = Font(
        name="Calibri",
        size=10,
        bold=False,
        italic=False,
        vertAlign=None,  # 첨자
        underline="none",  # 밑줄
        strike=False,  # 취소선
        color="00000000",  # 블랙, # 00FF0000 Red, # 000000FF Blue
    )
    wb = load_workbook(filename)
    ws = wb.sheetnames
    for sheet in ws:
        col_max = wb[sheet].max_column
        row_max = wb[sheet].max_row
        text_area.insert(tk.END, f"{sheet:<30}| Col = {col_max:<5}, Row = {row_max:<5} | ")
        text_area.see(tk.END)
        for row_c in range(i, row_max + 1, 1):
            for col_c in range(j, col_max + 1, 1):
                wb[sheet].cell(row=row_c, column=col_c).font = font_style
                wb[sheet].cell(row=row_c, column=col_c).alignment = Alignment(horizontal="right")
                # wb[sheet].cell(row=row_c, column=col_c).number_format = "#,##0.0"
                wb[sheet].cell(row=row_c, column=col_c).number_format = builtin_format_code(k)
        text_area.insert(tk.END, f"Done\n")
        text_area.see(tk.END)
    wb.save(filename)


def Common_daseul_log(list_file):
    list_file.delete(0, tk.END)
    files = filedialog.askopenfilenames(
        title="Cal log 파일을 선택하세요",
        filetypes=(("CSV 파일", "*.csv"), ("모든 파일", "*.*")),
        initialdir=r"C:\DGS\LOGS",
    )
    # 사용자가 선택한 파일 목록
    for file in files:
        list_file.insert(tk.END, file)


def Common_mtm_log(list_file):
    list_file.delete(0, tk.END)
    files = filedialog.askopenfilenames(
        title="MTM log 파일을 선택하세요",
        filetypes=(("CSV 파일", "*.csv"), ("모든 파일", "*.*")),
        initialdir=(r"D:\\DATA\\Project_DATA\\@_S23_FE\\TOOL\\2_MTM Calibration\\SM-S711B_MTM3_3.76.02_20230411_4Rx_Enable"),
    )
    # 사용자가 선택한 파일 목록
    for file in files:
        list_file.insert(tk.END, file)


def browse_spc_path(path_spc_file, daseul_select, text_area):
    spc_file_name = filedialog.askopenfilename(
        title="SPC 파일을 선택하세요",
        filetypes=(("SPC 파일", "*.dec"), ("모든 파일", "*.*")),
        initialdir=r"D:\\DATA\\TOOLS\\@_Python\\Auto_Cal_Spec",
    )

    if spc_file_name == "":  # 사용자가 취소를 누를 때
        text_area.insert(tk.END, "폴더 선택 취소\n")
        text_area.see(tk.END)
        return
    # print(folder_selected)
    path_spc_file.delete(0, tk.END)
    path_spc_file.insert(0, spc_file_name)
    daseul_select.set(True)


def browse_mtm_path(mtm_spc_file, mtm_select, text_area):
    mtm_folder_name = filedialog.askdirectory(
        title="MTM 폴더를 선택하세요",
        initialdir=r"D:\\DATA\\Project_DATA\\@_S23_FE\\TOOL\\2_MTM Calibration\\SM-S711B_MTM3_3.76.02_20230411_4Rx_Enable",
    )

    if mtm_folder_name == "":  # 사용자가 취소를 누를 때
        text_area.insert(tk.END, "폴더 선택 취소\n")
        text_area.see(tk.END)
        return
    # print(folder_selected)
    mtm_spc_file.delete(0, tk.END)
    mtm_spc_file.insert(0, mtm_folder_name)
    mtm_select.set(True)


def NR_channel_converter(band, channel):
    Info_freq_calc = {
        # fdl_low, nref_offset, ful_low, delta_f, freq_offset
        1: [2110, 0, 1920, 5, 0],
        2: [1930, 0, 1850, 5, 0],
        3: [1805, 0, 1710, 5, 0],
        5: [869, 0, 824, 5, 0],
        7: [2620, 0, 2500, 5, 0],
        8: [925, 0, 880, 5, 0],
        12: [729, 0, 699, 5, 0],
        13: [746, 0, 777, 5, 0],
        18: [860, 0, 815, 5, 0],
        20: [791, 0, 832, 5, 0],
        25: [1930, 0, 1850, 5, 0],
        26: [859, 0, 814, 5, 0],
        28: [758, 0, 703, 5, 0],
        66: [2110, 0, 1710, 5, 0],
        38: [2570, 0, 2570, 5, 0],
        39: [1880, 0, 1880, 5, 0],
        40: [2300, 0, 2300, 5, 0],
        41: [2496, 0, 2496, 5, 0],
        77: [3300, 600000, 3300, 15, 3000],
        78: [3300, 600000, 3300, 15, 3000],
    }

    freq_separation = {
        1: [190],
        2: [80],
        3: [95],
        5: [45],
        7: [120],
        8: [45],
        12: [30],
        13: [-31],
        18: [45],
        20: [-41],
        25: [80],
        26: [45],
        28: [55],
        66: [400],
        38: [0],
        39: [0],
        40: [0],
        41: [0],
        77: [0],
        78: [0],
    }

    # rxfreq = fref_offset + delta_f * (nref - nref_offset)
    # txfreq = ful_low, + 0.1*(nul - nul_offset)
    delta_f = Info_freq_calc[band][3] * 1000
    freq_offset = Info_freq_calc[band][4] * 1000000
    rx = int((freq_offset + (delta_f * (channel - Info_freq_calc[band][1]))) / 1000)
    tx = int((rx / 1000 - freq_separation[band][0]) * 1000)

    return rx, tx


def LTE_channel_converter(band, channel):
    Info_freq_calc = {
        # fdl_low, ndl_offset, ful_low, nul_offset
        1: [2110, 0, 1920, 18000],
        2: [1930, 600, 1850, 18600],
        3: [1805, 1200, 1710, 19200],
        4: [2110, 1950, 1710, 19950],
        5: [869, 2400, 824, 20400],
        7: [2620, 2750, 2500, 20750],
        8: [925, 3450, 880, 21450],
        12: [729, 5010, 699, 23010],
        13: [746, 5180, 777, 23180],
        17: [734, 5730, 704, 23730],
        18: [860, 5850, 815, 23850],
        19: [875, 6000, 830, 24000],
        20: [791, 6150, 832, 24150],
        25: [1930, 8040, 1850, 26040],
        26: [859, 8690, 814, 26690],
        28: [758, 9210, 703, 27210],
        66: [2110, 66436, 1710, 131972],
        38: [2570, 37750, 2570, 37750],
        39: [1880, 38250, 1880, 38250],
        40: [2300, 38650, 2300, 38650],
        41: [2496, 39650, 2496, 39650],
    }

    freq_separation = {
        1: [190],
        2: [80],
        3: [95],
        4: [400],
        5: [45],
        7: [120],
        8: [45],
        12: [30],
        13: [-31],
        17: [30],
        18: [45],
        19: [45],
        20: [-41],
        25: [80],
        26: [45],
        28: [55],
        66: [400],
        38: [0],
        39: [0],
        40: [0],
        41: [0],
    }

    # rxfreq = fdl_low, + 0.1*(ndl - ndl_offset)
    # txfreq = ful_low, + 0.1*(nul - nul_offset)
    rx = int((Info_freq_calc[band][0] + 0.1 * (channel - Info_freq_calc[band][1])) * 1000)
    tx = int((rx / 1000 - freq_separation[band][0]) * 1000)

    # tx = rx - np.float_(freq_separation[band][0])
    # tx = Info_freq_calc[band][2] + 0.1 * (channel - Info_freq_calc[band][3])

    return rx, tx


def HSPA_channel_converter(band, channel):
    Info_freq_calc = {
        # ndl_low, fdl_offset, nul_low, ful_offset
        1: [10562, 0, 9612, 0],
        2: [9662, 0, 9262, 0],
        4: [1537, 1805, 1312, 1805],
        5: [4357, 0, 4132, 0],
        8: [2937, 340, 2917, 340],
    }

    freq_separation = {
        1: [190],
        2: [80],
        4: [400],
        5: [45],
        8: [45],
    }

    # rxfreq = Fdl_offset + Ndl*0.2
    # txfreq = Ful_offset + Nul*0.2
    rx = int((channel * 0.2 + Info_freq_calc[band][1]) * 1000)
    tx = int((rx / 1000 - freq_separation[band][0]) * 1000)

    # tx = rx - np.float_(freq_separation[band][0])
    # tx = Info_freq_calc[band][2] + 0.1 * (channel - Info_freq_calc[band][3])

    return rx, tx
